"""文本提取与切块"""

import logging
import os
from typing import List, Optional

from src.models import Chunk

logger = logging.getLogger(__name__)


class TextChunker:
    """文本提取与切块器

    负责从多种格式的文档中提取纯文本，并将文本切分为适合向量化的块。
    """

    # 支持的文件扩展名到提取方法的映射
    SUPPORTED_EXTENSIONS = {
        ".txt": "_extract_txt",
        ".md": "_extract_markdown",
        ".doc": "_extract_word",
        ".docx": "_extract_word",
        ".xls": "_extract_excel",
        ".xlsx": "_extract_excel",
        ".pdf": "_extract_pdf",
    }

    def __init__(self, tokenizer=None, chunk_size: int = 512, chunk_overlap: int = 64):
        """初始化 TextChunker

        Args:
            tokenizer: tokenizer 实例，供切块使用（可选，文本提取不需要）
            chunk_size: 目标块大小（token），默认 512
            chunk_overlap: 重叠区域大小（token），默认 64
        """
        self.tokenizer = tokenizer
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    # 中文句子结束标点
    _SENTENCE_BOUNDARIES = set("。！？；\n")

    def chunk(self, text: str) -> List[Chunk]:
        """将文本切分为多个 Chunk

        使用 tokenizer 进行 token 计数，按 chunk_size 切分，
        相邻块保留 chunk_overlap token 重叠。
        以中文标点为优先分割边界。
        最后一块不足 chunk_overlap token 时合并到前一块。

        Args:
            text: 原始文本

        Returns:
            切块列表
        """
        if not self.tokenizer:
            raise ValueError("tokenizer 未设置，无法进行切块")

        if not text or not text.strip():
            return []

        # 对整个文本进行 tokenize（不添加特殊 token）
        token_ids = self.tokenizer.encode(text, add_special_tokens=False)
        total_tokens = len(token_ids)

        if total_tokens == 0:
            return []

        # 如果总 token 数不超过 chunk_size，直接返回单个 chunk
        if total_tokens <= self.chunk_size:
            return [Chunk(text=text.strip(), index=0, token_count=total_tokens)]

        chunks: List[Chunk] = []
        start = 0
        chunk_index = 0

        while start < total_tokens:
            end = min(start + self.chunk_size, total_tokens)

            if end < total_tokens:
                # 尝试在句子边界处切分
                boundary = self._find_sentence_boundary(token_ids, start, end)
                if boundary > start:
                    end = boundary

            chunk_token_ids = token_ids[start:end]
            chunk_text = self.tokenizer.decode(chunk_token_ids, skip_special_tokens=True)
            chunk_token_count = len(chunk_token_ids)

            # 检查是否是最后一块且不足 overlap token
            remaining = total_tokens - end
            if remaining > 0 and remaining < self.chunk_overlap:
                # 将剩余部分合并到当前块
                chunk_token_ids = token_ids[start:]
                chunk_text = self.tokenizer.decode(chunk_token_ids, skip_special_tokens=True)
                chunk_token_count = len(chunk_token_ids)
                chunks.append(Chunk(
                    text=chunk_text.strip(),
                    index=chunk_index,
                    token_count=chunk_token_count,
                ))
                break

            chunks.append(Chunk(
                text=chunk_text.strip(),
                index=chunk_index,
                token_count=chunk_token_count,
            ))

            if end >= total_tokens:
                break

            # 下一块从 end - overlap 开始，保留重叠
            start = end - self.chunk_overlap
            chunk_index += 1

        # 最后检查：如果最后一块不足 overlap token，合并到前一块
        if len(chunks) > 1 and chunks[-1].token_count < self.chunk_overlap:
            last_chunk = chunks.pop()
            # 重新计算前一块：从前一块的起始到文本末尾
            prev_chunk = chunks[-1]
            # 需要重新 decode 合并后的 token 范围
            # 计算前一块的起始 token 位置
            # 从 chunks 列表反推起始位置
            merged_start = self._get_chunk_start(chunks, token_ids)
            merged_token_ids = token_ids[merged_start:]
            merged_text = self.tokenizer.decode(merged_token_ids, skip_special_tokens=True)
            chunks[-1] = Chunk(
                text=merged_text.strip(),
                index=prev_chunk.index,
                token_count=len(merged_token_ids),
            )

        return chunks

    def _get_chunk_start(self, chunks: List[Chunk], all_token_ids: list) -> int:
        """计算最后一个 chunk 的起始 token 位置

        通过从头累加各 chunk 的实际前进步长来推算。
        """
        if len(chunks) == 0:
            return 0
        if len(chunks) == 1:
            return 0

        # 第一个 chunk 从 0 开始
        # 后续每个 chunk 的 start = 前一个 chunk 的 end - overlap
        # 第一个 chunk 的 end = chunk[0].token_count
        # 但由于句子边界切分，实际 end 可能小于 chunk_size
        # 我们需要重新计算
        start = 0
        for i, chunk in enumerate(chunks[:-1]):
            end = start + chunk.token_count
            start = end - self.chunk_overlap
        return start

    def _find_sentence_boundary(self, token_ids: list, start: int, max_end: int) -> int:
        """在 token 序列中找到最近的中文句子边界

        从 max_end 位置向前搜索，找到最近的句子结束标点位置。
        如果在 start 到 max_end 范围内找不到句子边界，返回 max_end。

        Args:
            token_ids: 完整文本的 token ID 序列
            start: 当前块的起始位置
            max_end: 最大结束位置（不超过 chunk_size）

        Returns:
            切分位置（token 索引，不包含该位置）
        """
        # 从 max_end 向前搜索句子边界
        # 策略：逐步向前 decode 单个 token，检查是否包含句子结束标点
        best_boundary = max_end

        # 从 max_end - 1 向 start 方向搜索
        # 为了效率，我们限制搜索范围（不超过 chunk_size 的一半）
        search_start = max(start + self.chunk_overlap, start)  # 至少保留 overlap 大小

        for pos in range(max_end - 1, search_start - 1, -1):
            # decode 当前 token 看是否包含句子边界字符
            token_text = self.tokenizer.decode([token_ids[pos]], skip_special_tokens=True)
            for ch in token_text:
                if ch in self._SENTENCE_BOUNDARIES:
                    # 在这个 token 之后切分（包含这个标点）
                    return pos + 1

        # 没找到句子边界，使用 max_end
        return max_end

    def extract_text(self, file_path: str) -> Optional[str]:
        """从文件提取纯文本

        根据文件扩展名分发到对应的提取器。

        Args:
            file_path: 文件绝对路径

        Returns:
            提取的纯文本内容，如果文件为空、损坏或不支持则返回 None
        """
        ext = os.path.splitext(file_path)[1].lower()

        if ext not in self.SUPPORTED_EXTENSIONS:
            logger.warning("不支持的文件格式: %s", file_path)
            return None

        method_name = self.SUPPORTED_EXTENSIONS[ext]
        extract_method = getattr(self, method_name)

        try:
            text = extract_method(file_path)
        except Exception as e:
            logger.error("文件提取失败 [%s]: %s", file_path, str(e))
            return None

        # 检查提取结果是否为空白
        if text is None or text.strip() == "":
            logger.warning("文件内容为空或提取后为空白: %s", file_path)
            return None

        return text

    def _extract_txt(self, file_path: str) -> Optional[str]:
        """提取 TXT 文件文本 - 直接 UTF-8 读取

        Args:
            file_path: 文件绝对路径

        Returns:
            文件文本内容
        """
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    def _extract_markdown(self, file_path: str) -> Optional[str]:
        """提取 Markdown 文件文本 - 解析后提取纯文本，移除语法标记保留结构换行

        Args:
            file_path: 文件绝对路径

        Returns:
            纯文本内容
        """
        from markdown_it import MarkdownIt

        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        md = MarkdownIt()
        tokens = md.parse(content)

        text_parts = []
        for token in tokens:
            if token.type == "inline" and token.children:
                inline_text = self._extract_inline_text(token.children)
                if inline_text:
                    text_parts.append(inline_text)
            elif token.type == "fence":
                # 代码块：保留代码内容
                if token.content:
                    text_parts.append(token.content.rstrip("\n"))
            elif token.type in (
                "heading_open",
                "paragraph_open",
                "bullet_list_open",
                "ordered_list_open",
            ):
                # 结构元素开始，不需要额外处理
                pass

        return "\n".join(text_parts)

    def _extract_inline_text(self, children) -> str:
        """从 inline token 的 children 中提取纯文本

        Args:
            children: inline token 的子 token 列表

        Returns:
            拼接后的纯文本
        """
        parts = []
        for child in children:
            if child.type == "text":
                parts.append(child.content)
            elif child.type == "code_inline":
                parts.append(child.content)
            elif child.type == "softbreak":
                parts.append("\n")
            elif child.type == "hardbreak":
                parts.append("\n")
        return "".join(parts)

    def _extract_word(self, file_path: str) -> Optional[str]:
        """提取 Word (.doc/.docx) 文件文本 - 按段落顺序提取

        Args:
            file_path: 文件绝对路径

        Returns:
            按段落拼接的文本内容
        """
        from docx import Document

        doc = Document(file_path)
        paragraphs = []
        for para in doc.paragraphs:
            text = para.text
            if text:
                paragraphs.append(text)

        return "\n".join(paragraphs)

    def _extract_excel(self, file_path: str) -> Optional[str]:
        """提取 Excel (.xls/.xlsx) 文件文本 - 按工作表→行→单元格拼接

        单元格之间以制表符分隔，行之间以换行符分隔。

        Args:
            file_path: 文件绝对路径

        Returns:
            拼接后的文本内容
        """
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_text = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                cells = []
                for cell in row:
                    value = cell.value
                    if value is not None:
                        cells.append(str(value))
                    else:
                        cells.append("")
                all_text.append("\t".join(cells))

        wb.close()
        return "\n".join(all_text)

    def _extract_pdf(self, file_path: str) -> Optional[str]:
        """提取 PDF 文件文本 - 按页码顺序提取

        Args:
            file_path: 文件绝对路径

        Returns:
            按页拼接的文本内容
        """
        from PyPDF2 import PdfReader

        reader = PdfReader(file_path)
        pages_text = []

        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages_text.append(text)

        return "\n".join(pages_text)
